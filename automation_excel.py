import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

archivo_excel = pd.read_excel('supermarket_sales.xlsx')
#print(archivo_excel[['Gender', 'Product line', 'Total']])

tabla_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)


tabla_pivote.to_excel('sales_2021.xlsx', startrow=4, sheet_name='Report')

wb = load_workbook('sales_2021.xlsx')
pestaña = wb['Report']

min_col = wb.active.min_column
max_col = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row

#grafico
barchart = BarChart()
data = Reference(pestaña, min)